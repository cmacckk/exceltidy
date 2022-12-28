from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pandas import read_excel
import xlwings as xw


def openpyxl_get_datas_workbook_worksheet(filename, sheet_name=None, data_only=True, read_only=True):
    """ 获取Excel数据和WorkSheet WorkBook

        Args:
            filename [str]: Excel文件名
            sheet_name [str]: 工作表名, 如果未传入, 默认读取第一个工作表
            data_only [bool]: 是否仅将数据转化为公式, True代表公式会进行计算
            read_only [bool]: 是否仅可读, 如果为True, 读取数据速度会提升很多,但不可以进行清除worksheet数据等操作
        
        Returns:
            worksheet_datas [list[list, list, ...]]: 工作表数据
            workbook [Workbook]: 工作薄
            worksheet [Worksheet]: 工作表
        
    """
    workbook = load_workbook(filename, read_only=read_only, data_only=data_only)

    if sheet_name is None:
        worksheet = workbook[workbook.sheetnames[0]]
    else:
        worksheet = workbook[sheet_name]

    worksheet_values = worksheet.values
    worksheet_datas = [[x for x in y] for y in worksheet_values]
    
    return worksheet_datas, workbook, worksheet

def openpyxl_get_datas(worksheet: Worksheet):
    worksheet_values = worksheet.values
    worksheet_datas = [[x for x in y] for y in worksheet_values]
    
    return worksheet_datas

def openpyxl_replace_worksheet_data(origin_worksheet: Worksheet, to_be_replaced_worksheet: Worksheet):
    openpyxl_clean_worksheet_datas(to_be_replaced_worksheet)
    datas = openpyxl_get_datas(origin_worksheet)
    openpyxl_datas_to_worksheet(datas=datas, worksheet=to_be_replaced_worksheet)

def openpyxl_clean_worksheet_datas(worksheet: Worksheet):
    """ 删除工作表中的所有数据, 因为使用delete_rows函数, 不会打乱原有数据基础上的公式

        Args:
            worksheet [Worksheet]: 要清除所有数据的工作表
    """
    max_row = worksheet.max_row
    worksheet.delete_rows(1, max_row)

def openpyxl_datas_to_worksheet(datas: list, worksheet: Worksheet):
    """ 将读取出来的数据列表填入到工作表中

        Args:
            datas list[list, list, ...]: 要填入的数据
            worksheet [Worksheet]: 被数据填入的工作表
    """
    for row in range(len(datas)):
        for col in range(len(datas[row])):
            worksheet.cell(row=row + 1, column=col + 1).value = datas[row][col]

def xlwings_get_workbook_worksheet(filename, sheet_name=None, visible=False, add_book=False, display_alerts=False, screen_updating=False):
    app = xw.App(visible=visible, add_book=add_book)
    app.display_alerts = display_alerts
    app.screen_updating = screen_updating

    workbook = app.books.open(filename)

    if sheet_name is None:
        worksheet = workbook.sheets[0]
    else:
        worksheet = workbook.sheets[sheet_name]

    return workbook, worksheet

def xlwings_replace_worksheet_data(origin_worksheet: xw.Sheet, to_be_replaced_worksheet: xw.Sheet):
    origin_worksheet.api.copy()

if __name__ == "__main__":
    datas, workbook, worksheet = openpyxl_get_datas_workbook_worksheet('./收入积分_20221212-20221218.xlsx', data_only=False, read_only=True)
    new_datas, new_workbook, new_worksheet = openpyxl_get_datas_workbook_worksheet('./积分.xlsx', sheet_name="12月", data_only=False, read_only=False)
    openpyxl_replace_worksheet_data(worksheet, new_worksheet)
    new_workbook.save('./test.xlsx')